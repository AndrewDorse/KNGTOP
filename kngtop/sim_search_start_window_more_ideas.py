from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

WINDOW_SEC = 300
FILL_DELAY_SEC = 2
NOTIONAL_USD = 1.0
DEFAULT_SAMPLE_SIZES = ("100", "200", "400", "1000", "full")
_SLUG_EP = re.compile(r"-(\d+)$")


@dataclass(frozen=True, slots=True)
class Tick:
    elapsed_sec: int
    up_price: float
    down_price: float
    btc_price: float
    btc_volume: float


@dataclass(frozen=True, slots=True)
class WindowData:
    slug: str
    path: Path
    question: str
    start_ms: int
    ticks: tuple[Tick, ...]


@dataclass(frozen=True, slots=True)
class Variant:
    key: str
    family: str
    label: str
    price_min: float
    price_max: float
    entry_max_sec: int
    move_from_open_min_usd: float
    volume_ratio_min: float | None
    volume_lookback_sec: int
    pm_push_lookback_sec: int
    pm_push_cap: float | None
    pm_push_min: float | None
    reclaim_level: float | None


@dataclass(slots=True)
class SampleStats:
    windows: int
    trades: int = 0
    wins: int = 0
    total_pnl_usd: float = 0.0
    entry_sum: float = 0.0

    def record(self, *, traded: bool, pnl_usd: float, entry_px: float | None) -> None:
        if not traded or entry_px is None:
            return
        self.trades += 1
        self.total_pnl_usd += pnl_usd
        self.entry_sum += entry_px
        if pnl_usd > 1e-12:
            self.wins += 1

    @property
    def trade_rate_pct(self) -> float:
        return (100.0 * self.trades / self.windows) if self.windows else 0.0

    @property
    def win_rate_pct(self) -> float:
        return (100.0 * self.wins / self.trades) if self.trades else 0.0

    @property
    def avg_entry_px(self) -> float:
        return (self.entry_sum / self.trades) if self.trades else 0.0

    @property
    def avg_pnl_per_trade(self) -> float:
        return (self.total_pnl_usd / self.trades) if self.trades else 0.0

    @property
    def avg_pnl_per_window(self) -> float:
        return (self.total_pnl_usd / self.windows) if self.windows else 0.0


def slug_epoch(slug: str) -> int:
    m = _SLUG_EP.search((slug or "").strip())
    return int(m.group(1)) if m else 0


def load_windows(input_dir: Path) -> list[WindowData]:
    out: list[WindowData] = []
    for path in sorted(input_dir.glob("*_btc-updown-5m-*_prices.csv")):
        with path.open("r", encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle))
        if not rows:
            continue
        by_elapsed: dict[int, dict[str, str]] = {}
        for row in rows:
            try:
                elapsed = int(float(row.get("elapsed_sec") or 0))
            except (TypeError, ValueError):
                continue
            if 0 <= elapsed < WINDOW_SEC:
                by_elapsed[elapsed] = row
        if not by_elapsed:
            continue
        slug = str(rows[0].get("slug") or path.stem)
        question = str(rows[0].get("question") or "")
        last_up = 0.5
        last_dn = 0.5
        last_btc = 0.0
        ticks: list[Tick] = []
        for elapsed in range(WINDOW_SEC):
            row = by_elapsed.get(elapsed)
            if row is not None:
                try:
                    raw = row.get("up_price")
                    if raw not in (None, ""):
                        last_up = float(raw)
                    raw = row.get("down_price")
                    if raw not in (None, ""):
                        last_dn = float(raw)
                    raw = row.get("btc_price")
                    if raw not in (None, ""):
                        last_btc = float(raw)
                except (TypeError, ValueError):
                    pass
                try:
                    btc_volume = float(row.get("btc_volume") or 0.0)
                except (TypeError, ValueError):
                    btc_volume = 0.0
            else:
                btc_volume = 0.0
            ticks.append(
                Tick(
                    elapsed_sec=elapsed,
                    up_price=last_up,
                    down_price=last_dn,
                    btc_price=last_btc,
                    btc_volume=max(0.0, btc_volume),
                )
            )
        if ticks[0].btc_price <= 0 or any(t.btc_price <= 0 for t in ticks[:5]):
            continue
        out.append(
            WindowData(
                slug=slug,
                path=path,
                question=question,
                start_ms=slug_epoch(slug) * 1000,
                ticks=tuple(ticks),
            )
        )
    out.sort(key=lambda win: win.start_ms, reverse=True)
    return out


def winner_side(win: WindowData, tick_index: int) -> str | None:
    delta = win.ticks[tick_index].btc_price - win.ticks[0].btc_price
    if delta > 1e-12:
        return "UP"
    if delta < -1e-12:
        return "DOWN"
    return None


def side_price(tick: Tick, side: str) -> float:
    return tick.up_price if side == "UP" else tick.down_price


def volume_ratio(win: WindowData, tick_index: int, lookback_sec: int) -> float:
    current = win.ticks[tick_index].btc_volume
    lo = max(0, tick_index - int(lookback_sec))
    prev = [win.ticks[i].btc_volume for i in range(lo, tick_index)]
    if not prev:
        return 0.0
    mean_prev = sum(prev) / len(prev)
    if current <= 0.0 and mean_prev <= 0.0:
        return 0.0
    if mean_prev <= 0.0:
        return float("inf")
    return current / mean_prev


def side_price_push(win: WindowData, tick_index: int, side: str, lookback_sec: int) -> float:
    lo = max(0, tick_index - int(lookback_sec))
    return side_price(win.ticks[tick_index], side) - side_price(win.ticks[lo], side)


def delayed_fill(win: WindowData, tick_index: int, side: str) -> tuple[int, float] | None:
    fill_t = tick_index + FILL_DELAY_SEC
    if fill_t >= len(win.ticks):
        return None
    return fill_t, side_price(win.ticks[fill_t], side)


def settle(win: WindowData, side: str, entry_px: float) -> float:
    shares = NOTIONAL_USD / entry_px
    start_btc = win.ticks[0].btc_price
    end_btc = win.ticks[-1].btc_price
    if abs(end_btc - start_btc) < 1e-12:
        last_px = side_price(win.ticks[-1], side)
        return shares * last_px - NOTIONAL_USD
    won = (end_btc > start_btc and side == "UP") or (end_btc < start_btc and side == "DOWN")
    return (shares - NOTIONAL_USD) if won else (-NOTIONAL_USD)


def variant_passes(win: WindowData, tick_index: int, variant: Variant) -> tuple[bool, str | None]:
    side = winner_side(win, tick_index)
    if side is None:
        return False, None
    tick = win.ticks[tick_index]
    px = side_price(tick, side)
    if px < variant.price_min - 1e-12 or px > variant.price_max + 1e-12:
        return False, side
    if abs(tick.btc_price - win.ticks[0].btc_price) + 1e-12 < variant.move_from_open_min_usd:
        return False, side
    if variant.volume_ratio_min is not None:
        if volume_ratio(win, tick_index, variant.volume_lookback_sec) + 1e-12 < variant.volume_ratio_min:
            return False, side
    if variant.pm_push_cap is not None:
        if side_price_push(win, tick_index, side, variant.pm_push_lookback_sec) > variant.pm_push_cap + 1e-12:
            return False, side
    if variant.pm_push_min is not None:
        if side_price_push(win, tick_index, side, variant.pm_push_lookback_sec) + 1e-12 < variant.pm_push_min:
            return False, side
    if variant.reclaim_level is not None:
        if tick_index < 1:
            return False, side
        prev_px = side_price(win.ticks[tick_index - 1], side)
        if not (prev_px < variant.reclaim_level <= px):
            return False, side
    return True, side


def generate_variants() -> list[Variant]:
    out: list[Variant] = []
    seq = 1

    for price_min, price_max in ((0.44, 0.54), (0.45, 0.55)):
        for entry_max_sec in (8, 12, 20):
            for move_min in (1.0, 2.0, 4.0):
                for push_lookback_sec in (3, 5):
                    for push_cap in (0.00, 0.01, 0.02):
                        out.append(
                            Variant(
                                key=f"N{seq:04d}",
                                family="lag_follow",
                                label=(
                                    f"winner {price_min:.2f}-{price_max:.2f} first{entry_max_sec}s "
                                    f"move>={move_min:.1f} pm_push{push_lookback_sec}<={push_cap:.2f}"
                                ),
                                price_min=price_min,
                                price_max=price_max,
                                entry_max_sec=entry_max_sec,
                                move_from_open_min_usd=move_min,
                                volume_ratio_min=None,
                                volume_lookback_sec=0,
                                pm_push_lookback_sec=push_lookback_sec,
                                pm_push_cap=push_cap,
                                pm_push_min=None,
                                reclaim_level=None,
                            )
                        )
                        seq += 1
                        for vol_min in (1.5, 2.0):
                            out.append(
                                Variant(
                                    key=f"N{seq:04d}",
                                    family="lag_follow",
                                    label=(
                                        f"winner {price_min:.2f}-{price_max:.2f} first{entry_max_sec}s "
                                        f"move>={move_min:.1f} pm_push{push_lookback_sec}<={push_cap:.2f} "
                                        f"vol10>={vol_min:.1f}x"
                                    ),
                                    price_min=price_min,
                                    price_max=price_max,
                                    entry_max_sec=entry_max_sec,
                                    move_from_open_min_usd=move_min,
                                    volume_ratio_min=vol_min,
                                    volume_lookback_sec=10,
                                    pm_push_lookback_sec=push_lookback_sec,
                                    pm_push_cap=push_cap,
                                    pm_push_min=None,
                                    reclaim_level=None,
                                )
                            )
                            seq += 1

    for price_min, price_max in ((0.46, 0.56), (0.48, 0.58)):
        for entry_max_sec in (10, 15, 20):
            for move_min in (1.0, 2.0, 4.0):
                for push_lookback_sec in (5, 8):
                    for push_min in (0.02, 0.03):
                        out.append(
                            Variant(
                                key=f"N{seq:04d}",
                                family="accel_breakout",
                                label=(
                                    f"winner {price_min:.2f}-{price_max:.2f} first{entry_max_sec}s "
                                    f"move>={move_min:.1f} pm_push{push_lookback_sec}>={push_min:.2f}"
                                ),
                                price_min=price_min,
                                price_max=price_max,
                                entry_max_sec=entry_max_sec,
                                move_from_open_min_usd=move_min,
                                volume_ratio_min=None,
                                volume_lookback_sec=0,
                                pm_push_lookback_sec=push_lookback_sec,
                                pm_push_cap=None,
                                pm_push_min=push_min,
                                reclaim_level=None,
                            )
                        )
                        seq += 1
                        out.append(
                            Variant(
                                key=f"N{seq:04d}",
                                family="accel_breakout",
                                label=(
                                    f"winner {price_min:.2f}-{price_max:.2f} first{entry_max_sec}s "
                                    f"move>={move_min:.1f} pm_push{push_lookback_sec}>={push_min:.2f} "
                                    f"vol10>=1.5x"
                                ),
                                price_min=price_min,
                                price_max=price_max,
                                entry_max_sec=entry_max_sec,
                                move_from_open_min_usd=move_min,
                                volume_ratio_min=1.5,
                                volume_lookback_sec=10,
                                pm_push_lookback_sec=push_lookback_sec,
                                pm_push_cap=None,
                                pm_push_min=push_min,
                                reclaim_level=None,
                            )
                        )
                        seq += 1

    for price_min, price_max in ((0.48, 0.58), (0.50, 0.60)):
        for entry_max_sec in (15, 20, 30):
            for move_min in (2.0, 4.0, 6.0):
                for reclaim_level in (0.50, 0.52):
                    out.append(
                        Variant(
                            key=f"N{seq:04d}",
                            family="midline_reclaim",
                            label=(
                                f"winner {price_min:.2f}-{price_max:.2f} first{entry_max_sec}s "
                                f"cross>{reclaim_level:.2f} move>={move_min:.1f}"
                            ),
                            price_min=price_min,
                            price_max=price_max,
                            entry_max_sec=entry_max_sec,
                            move_from_open_min_usd=move_min,
                            volume_ratio_min=None,
                            volume_lookback_sec=0,
                            pm_push_lookback_sec=1,
                            pm_push_cap=None,
                            pm_push_min=None,
                            reclaim_level=reclaim_level,
                        )
                    )
                    seq += 1
                    out.append(
                        Variant(
                            key=f"N{seq:04d}",
                            family="midline_reclaim",
                            label=(
                                f"winner {price_min:.2f}-{price_max:.2f} first{entry_max_sec}s "
                                f"cross>{reclaim_level:.2f} move>={move_min:.1f} vol10>=1.5x"
                            ),
                            price_min=price_min,
                            price_max=price_max,
                            entry_max_sec=entry_max_sec,
                            move_from_open_min_usd=move_min,
                            volume_ratio_min=1.5,
                            volume_lookback_sec=10,
                            pm_push_lookback_sec=1,
                            pm_push_cap=None,
                            pm_push_min=None,
                            reclaim_level=reclaim_level,
                        )
                    )
                    seq += 1

    return out


def parse_sample_sizes(raw: str, full_n: int) -> list[tuple[str, int]]:
    out: list[tuple[str, int]] = []
    for token in (part.strip().lower() for part in raw.split(",") if part.strip()):
        if token == "full":
            out.append(("full", full_n))
        else:
            n = int(token)
            if n <= full_n:
                out.append((str(n), n))
    if not out:
        out = [(label, full_n if label == "full" else int(label)) for label in DEFAULT_SAMPLE_SIZES if label == "full" or int(label) <= full_n]
    seen: set[str] = set()
    deduped: list[tuple[str, int]] = []
    for label, n in out:
        if label in seen:
            continue
        seen.add(label)
        deduped.append((label, n))
    return deduped


def gate_thresholds(label: str, full_n: int) -> tuple[int, float]:
    if label == "100":
        return 10, 55.0
    if label == "200":
        return 20, 55.0
    if label == "400":
        return 40, 55.0
    if label == "1000":
        return 80, 55.0
    if label == "full":
        return max(200, min(800, full_n // 3)), 55.0
    return 0, 0.0


def evaluate_variant(
    windows: list[WindowData],
    variant: Variant,
    sample_specs: list[tuple[str, int]],
) -> dict[str, object]:
    stats = {label: SampleStats(windows=n) for label, n in sample_specs}
    for window_index, win in enumerate(windows):
        traded = False
        entry_px: float | None = None
        pnl_usd = 0.0
        for tick_index, tick in enumerate(win.ticks):
            if tick.elapsed_sec > variant.entry_max_sec:
                break
            passed, side = variant_passes(win, tick_index, variant)
            if not passed or side is None:
                continue
            fill = delayed_fill(win, tick_index, side)
            if fill is None:
                break
            _, fill_px = fill
            if fill_px < variant.price_min - 1e-12 or fill_px > variant.price_max + 1e-12:
                break
            traded = True
            entry_px = fill_px
            pnl_usd = settle(win, side, fill_px)
            break
        for label, n in sample_specs:
            if window_index < n:
                stats[label].record(traded=traded, pnl_usd=pnl_usd, entry_px=entry_px)

    row: dict[str, object] = {
        "variant_key": variant.key,
        "family": variant.family,
        "variant_label": variant.label,
        "notes": (
            "Round-2 start-window search. Single $1 buy on one side only, fill booked at t+2 from the "
            "public snapshot, then hold to 5m expiry."
        ),
    }
    for label, n in sample_specs:
        sample = stats[label]
        row[f"windows_{label}"] = n
        row[f"trades_{label}"] = sample.trades
        row[f"trade_rate_pct_{label}"] = round(sample.trade_rate_pct, 4)
        row[f"win_rate_pct_{label}"] = round(sample.win_rate_pct, 4)
        row[f"total_pnl_usd_{label}"] = round(sample.total_pnl_usd, 6)
        row[f"avg_pnl_per_trade_{label}"] = round(sample.avg_pnl_per_trade, 6)
        row[f"avg_pnl_per_window_{label}"] = round(sample.avg_pnl_per_window, 6)
        row[f"avg_entry_px_{label}"] = round(sample.avg_entry_px, 6)
        min_trades, min_wr = gate_thresholds(label, len(windows))
        row[f"passes_gate_{label}"] = (
            sample.trades >= min_trades
            and sample.win_rate_pct + 1e-12 >= min_wr
            and sample.total_pnl_usd > 0.0
        )
    gate_labels = [label for label, _ in sample_specs if label in {"100", "200", "400", "1000"}]
    row["passes_progressive_1000_gate"] = all(bool(row[f"passes_gate_{label}"]) for label in gate_labels)
    row["passes_full_sanity"] = bool(row.get("passes_progressive_1000_gate")) and bool(row.get("passes_gate_full", False))
    return row


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx_if_available(
    summary_rows: list[dict[str, object]],
    best_rows: list[dict[str, object]],
    family_rows: list[dict[str, object]],
    output_csv: Path,
) -> None:
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "best_realistic"
    ws.append(list(best_rows[0].keys()))
    for row in best_rows:
        ws.append([row[key] for key in best_rows[0].keys()])
    ws2 = wb.create_sheet("family_leaders")
    ws2.append(list(family_rows[0].keys()))
    for row in family_rows:
        ws2.append([row[key] for key in family_rows[0].keys()])
    ws3 = wb.create_sheet("summary_all")
    ws3.append(list(summary_rows[0].keys()))
    for row in summary_rows:
        ws3.append([row[key] for key in summary_rows[0].keys()])
    wb.save(output_csv.with_suffix(".xlsx"))


def build_family_leaders(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    leaders: dict[str, dict[str, object]] = {}
    for row in rows:
        family = str(row["family"])
        current = leaders.get(family)
        score = (
            bool(row["passes_full_sanity"]),
            bool(row["passes_progressive_1000_gate"]),
            float(row.get("total_pnl_usd_1000", 0.0)),
            float(row.get("win_rate_pct_1000", 0.0)),
            float(row.get("total_pnl_usd_full", 0.0)),
        )
        if current is None:
            leaders[family] = row
            continue
        current_score = (
            bool(current["passes_full_sanity"]),
            bool(current["passes_progressive_1000_gate"]),
            float(current.get("total_pnl_usd_1000", 0.0)),
            float(current.get("win_rate_pct_1000", 0.0)),
            float(current.get("total_pnl_usd_full", 0.0)),
        )
        if score > current_score:
            leaders[family] = row
    return sorted(
        leaders.values(),
        key=lambda row: (
            bool(row["passes_full_sanity"]),
            bool(row["passes_progressive_1000_gate"]),
            float(row.get("total_pnl_usd_1000", 0.0)),
        ),
        reverse=True,
    )


def run_search(
    input_dir: Path,
    out_summary: Path,
    out_best: Path,
    out_family: Path,
    sample_sizes_raw: str,
) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    windows = load_windows(input_dir)
    if not windows:
        raise RuntimeError(f"No usable windows in {input_dir}")
    sample_specs = parse_sample_sizes(sample_sizes_raw, len(windows))
    variants = generate_variants()
    summary_rows = [evaluate_variant(windows, variant, sample_specs) for variant in variants]
    summary_rows.sort(
        key=lambda row: (
            bool(row["passes_full_sanity"]),
            bool(row["passes_progressive_1000_gate"]),
            float(row.get("total_pnl_usd_1000", 0.0)),
            float(row.get("win_rate_pct_1000", 0.0)),
            float(row.get("total_pnl_usd_full", 0.0)),
        ),
        reverse=True,
    )
    best_rows = [row for row in summary_rows if bool(row["passes_full_sanity"])]
    family_rows = build_family_leaders(summary_rows)
    write_csv(out_summary, summary_rows)
    write_csv(out_best, best_rows)
    write_csv(out_family, family_rows)
    write_xlsx_if_available(summary_rows, best_rows, family_rows, out_best)
    return summary_rows, best_rows, family_rows


def print_rankings(rows: list[dict[str, object]], *, title: str, limit: int = 12) -> None:
    print(title)
    for row in rows[:limit]:
        print(
            f"  {row['variant_key']} | {row['family']} | "
            f"pnl1000={float(row.get('total_pnl_usd_1000', 0.0)):+.2f} "
            f"wr1000={float(row.get('win_rate_pct_1000', 0.0)):.2f}% "
            f"trades1000={int(row.get('trades_1000', 0))} "
            f"full_ok={bool(row.get('passes_full_sanity'))} | {row['variant_label']}"
        )
    print()


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    git_root = repo_root.parent
    parser = argparse.ArgumentParser(description="Round-2 BTC 5m start-window idea search.")
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=git_root / "kng_bot3" / "exports" / "window_price_snapshots_public" / "btc_5m",
    )
    parser.add_argument(
        "--out-summary",
        type=Path,
        default=repo_root / "reports" / "start_window_more_ideas_summary.csv",
    )
    parser.add_argument(
        "--out-best",
        type=Path,
        default=repo_root / "reports" / "start_window_more_ideas_best_realistic.csv",
    )
    parser.add_argument(
        "--out-family",
        type=Path,
        default=repo_root / "reports" / "start_window_more_ideas_family_leaders.csv",
    )
    parser.add_argument("--sample-sizes", default="100,200,400,1000,full")
    args = parser.parse_args()

    summary_rows, best_rows, family_rows = run_search(
        args.input_dir,
        args.out_summary,
        args.out_best,
        args.out_family,
        args.sample_sizes,
    )
    print_rankings(best_rows, title="Best realistic variants:")
    print_rankings(family_rows, title="Best family leaders:", limit=len(family_rows))


if __name__ == "__main__":
    main()
